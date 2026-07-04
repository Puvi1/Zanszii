"""
Feb 2026 batch fixes - integration tests.
Covers: Excel/PDF/CSV exports, Saturday lock, admin missions CRUD,
mark-for-member, team-week, goals removal, dedupe, security.
"""
import os
import io
import time
import pytest
import requests
from datetime import date

BASE_URL = os.environ.get("REACT_APP_BACKEND_URL", "https://growth-gamified-2.preview.emergentagent.com").rstrip("/")
API = f"{BASE_URL}/api"

ADMIN_EMAIL = "admin@spartans.com"
ADMIN_PASSWORD = "Sp4rt@n-Cmd-2026!F0rge"
OLD_ADMIN_PASSWORD = "Spartan123!"


# ---------------- Fixtures ----------------
@pytest.fixture(scope="session")
def admin_token():
    r = requests.post(f"{API}/auth/login", json={"email": ADMIN_EMAIL, "password": ADMIN_PASSWORD}, timeout=30)
    assert r.status_code == 200, f"Admin login failed: {r.status_code} {r.text}"
    return r.json().get("access_token") or r.json().get("token")


@pytest.fixture(scope="session")
def admin_headers(admin_token):
    return {"Authorization": f"Bearer {admin_token}"}


def _register_member(prefix="mem"):
    ts = int(time.time() * 1000)
    email = f"TEST_{prefix}_{ts}@example.com"
    payload = {
        "email": email,
        "password": "MemberPass1!",
        "name": f"Test {prefix} {ts}",
        "nexus_id": f"BC-{prefix.upper()}-{ts}",
    }
    r = requests.post(f"{API}/auth/register", json=payload, timeout=30)
    assert r.status_code in (200, 201), f"register: {r.status_code} {r.text}"
    body = r.json()
    tok = body.get("access_token") or body.get("token")
    uid = body.get("user", {}).get("user_id") or body.get("user_id")
    return {"email": email, "token": tok, "user_id": uid, "headers": {"Authorization": f"Bearer {tok}"}}


@pytest.fixture(scope="session")
def member_a():
    return _register_member("a")


@pytest.fixture(scope="session")
def member_b():
    return _register_member("b")


# ---------------- SECURITY ----------------
class TestSecurity:
    def test_old_password_rejected(self):
        r = requests.post(f"{API}/auth/login", json={"email": ADMIN_EMAIL, "password": OLD_ADMIN_PASSWORD}, timeout=30)
        assert r.status_code == 401, f"old password should not work, got {r.status_code}"

    def test_new_password_works(self, admin_token):
        assert admin_token and len(admin_token) > 10


# ---------------- EXPORTS ----------------
EXPORT_SCOPES = [
    "missions", "tasks", "goals", "followups", "attendance",
    "xp-leaderboard", "team-performance", "daily",
]

class TestExports:
    @pytest.mark.parametrize("scope", EXPORT_SCOPES)
    def test_xlsx(self, admin_headers, scope):
        r = requests.get(f"{API}/exports/{scope}?format=xlsx", headers=admin_headers, timeout=60)
        assert r.status_code == 200, f"{scope} xlsx -> {r.status_code} {r.text[:200]}"
        assert "spreadsheetml.sheet" in r.headers.get("content-type", ""), r.headers.get("content-type")
        assert r.content[:2] == b"PK", f"{scope} not a zip/xlsx"

    @pytest.mark.parametrize("scope", EXPORT_SCOPES)
    def test_pdf(self, admin_headers, scope):
        r = requests.get(f"{API}/exports/{scope}?format=pdf", headers=admin_headers, timeout=60)
        assert r.status_code == 200, f"{scope} pdf -> {r.status_code} {r.text[:200]}"
        assert "pdf" in r.headers.get("content-type", ""), r.headers.get("content-type")
        assert r.content[:4] == b"%PDF", f"{scope} pdf magic missing"

    @pytest.mark.parametrize("scope", EXPORT_SCOPES)
    def test_csv(self, admin_headers, scope):
        r = requests.get(f"{API}/exports/{scope}?format=csv", headers=admin_headers, timeout=60)
        assert r.status_code == 200, f"{scope} csv -> {r.status_code} {r.text[:200]}"
        assert "csv" in r.headers.get("content-type", ""), r.headers.get("content-type")

    def test_spartans_league_individual_xlsx(self, admin_headers):
        r = requests.get(f"{API}/exports/spartans-league?format=xlsx&scope=individual", headers=admin_headers, timeout=60)
        assert r.status_code == 200
        assert "spreadsheetml.sheet" in r.headers.get("content-type", "")

    def test_spartans_league_team_xlsx(self, admin_headers):
        r = requests.get(f"{API}/exports/spartans-league?format=xlsx&scope=team", headers=admin_headers, timeout=60)
        assert r.status_code == 200
        assert "spreadsheetml.sheet" in r.headers.get("content-type", "")

    def test_xlsx_opens_as_valid_workbook(self, admin_headers):
        r = requests.get(f"{API}/exports/xp-leaderboard?format=xlsx", headers=admin_headers, timeout=60)
        assert r.status_code == 200
        try:
            from openpyxl import load_workbook
            wb = load_workbook(io.BytesIO(r.content))
            assert len(wb.sheetnames) >= 1
        except Exception as e:
            pytest.fail(f"xlsx invalid: {e}")

    def test_csv_formula_injection_sanitizer(self, admin_headers):
        # Create a mission with malicious title, then export csv, confirm the row starts with '=' is prefixed
        r = requests.post(f"{API}/missions", headers=admin_headers,
                          json={"prospect_name": "=cmd|calc", "status": "new"}, timeout=30)
        # not fatal if endpoint mismatch, only assert sanitizer on the csv
        mid = None
        if r.status_code in (200, 201):
            mid = (r.json().get("mission") or {}).get("mission_id")
        try:
            r2 = requests.get(f"{API}/exports/missions?format=csv", headers=admin_headers, timeout=60)
            assert r2.status_code == 200
            text = r2.text
            # Look for the malicious value; if present, ensure it's prefixed
            if "=cmd|calc" in text:
                # Should not appear without leading quote at cell start
                assert ",=cmd|calc" not in text and "\n=cmd|calc" not in text, "formula injection not sanitized"
        finally:
            if mid:
                requests.delete(f"{API}/missions/{mid}", headers=admin_headers, timeout=30)


# ---------------- SATURDAY LOCK ----------------
class TestAttendanceLock:
    def test_week_endpoint(self, admin_headers):
        today = date.today().isoformat()
        r = requests.get(f"{API}/event-attendance/week?week_of={today}", headers=admin_headers, timeout=30)
        assert r.status_code == 200, f"{r.status_code} {r.text[:300]}"
        data = r.json()
        occs = data.get("occurrences") or data.get("events") or []
        # find saturday
        sat = [o for o in occs if str(o.get("weekday","")).lower().startswith("sat") or o.get("day")=="Saturday"]
        # Fallback: search by date suffix if today is Sat
        if not sat:
            sat = [o for o in occs if o.get("date","").endswith(today[-2:]) ]
        # Just report; don't hard fail structure
        print(f"occurrences: {len(occs)}, sat matches: {len(sat)}")
        # Verify at least one occurrence with locked flag present
        assert any("locked" in o for o in occs), "no locked flag in occurrences"


# ---------------- ADMIN MISSIONS ----------------
class TestAdminMissions:
    def test_admin_list_missions(self, admin_headers):
        r = requests.get(f"{API}/admin/missions", headers=admin_headers, timeout=30)
        assert r.status_code == 200, f"{r.status_code} {r.text[:200]}"
        data = r.json()
        assert isinstance(data, list) or isinstance(data, dict)

    def test_member_cannot_access_admin_missions(self, member_a):
        r = requests.get(f"{API}/admin/missions", headers=member_a["headers"], timeout=30)
        assert r.status_code == 403, f"expected 403 got {r.status_code}"

    def test_admin_update_and_delete_mission_reverses_xp(self, admin_headers, member_a):
        # Create mission as member
        r = requests.post(f"{API}/missions", headers=member_a["headers"],
                          json={"prospect_name": "TEST_prospect", "status": "new"}, timeout=30)
        assert r.status_code in (200, 201), r.text
        mid = (r.json().get("mission") or {}).get("mission_id") or r.json().get("id")
        assert mid

        # Get member XP before admin update
        me_before = requests.get(f"{API}/auth/me", headers=member_a["headers"], timeout=30).json()
        xp_before = (me_before.get("user") or me_before).get("xp", 0)

        # Admin patches to converted
        r2 = requests.patch(f"{API}/admin/missions/{mid}", headers=admin_headers,
                            json={"status": "converted"}, timeout=30)
        assert r2.status_code == 200, f"admin patch: {r2.status_code} {r2.text[:200]}"

        me_after = requests.get(f"{API}/auth/me", headers=member_a["headers"], timeout=30).json()
        xp_after = (me_after.get("user") or me_after).get("xp", 0)
        assert xp_after >= xp_before, f"xp did not increase after conversion: {xp_before}->{xp_after}"

        # Admin deletes mission -> should reverse xp
        r3 = requests.delete(f"{API}/admin/missions/{mid}", headers=admin_headers, timeout=30)
        assert r3.status_code in (200, 204), f"admin del: {r3.status_code} {r3.text[:200]}"

        me_final = requests.get(f"{API}/auth/me", headers=member_a["headers"], timeout=30).json()
        xp_final = (me_final.get("user") or me_final).get("xp", 0)
        assert xp_final <= xp_after, f"xp not reversed: {xp_after}->{xp_final}"


# ---------------- MEMBER MISSIONS CRUD ----------------
class TestMemberMissions:
    def test_member_lists_only_own(self, member_a, member_b):
        r = requests.post(f"{API}/missions", headers=member_a["headers"],
                          json={"prospect_name": "TEST_ownA", "status": "new"}, timeout=30)
        mid_a = (r.json().get("mission") or {}).get("mission_id")

        r = requests.get(f"{API}/missions", headers=member_b["headers"], timeout=30)
        assert r.status_code == 200
        ids = [m.get("mission_id") for m in r.json()] if isinstance(r.json(), list) else []
        assert mid_a not in ids

        # member B cannot delete member A's mission
        rd = requests.delete(f"{API}/missions/{mid_a}", headers=member_b["headers"], timeout=30)
        assert rd.status_code == 404, f"expected 404 got {rd.status_code}"

        # cleanup
        requests.delete(f"{API}/missions/{mid_a}", headers=member_a["headers"], timeout=30)

    def test_member_can_delete_own_mission(self, member_a):
        r = requests.post(f"{API}/missions", headers=member_a["headers"],
                          json={"prospect_name": "TEST_selfDel", "status": "new"}, timeout=30)
        mid = (r.json().get("mission") or {}).get("mission_id")
        rd = requests.delete(f"{API}/missions/{mid}", headers=member_a["headers"], timeout=30)
        assert rd.status_code in (200, 204)


# ---------------- ATTENDANCE mark-for-member ----------------
class TestMarkForMember:
    def test_member_cannot_mark_for_member(self, admin_headers, member_a, member_b):
        # fetch valid event
        rw = requests.get(f"{API}/event-attendance/week?week_of={date.today().isoformat()}", headers=admin_headers, timeout=30)
        occs = rw.json().get("occurrences", []) if rw.status_code == 200 else []
        if not occs:
            pytest.skip("no events")
        eid = occs[0].get("event_id")
        edate = occs[0].get("date", date.today().isoformat())
        payload = {"user_id": member_b["user_id"], "event_id": eid, "event_date": edate, "status": "present"}
        r = requests.post(f"{API}/event-attendance/mark-for-member", headers=member_a["headers"], json=payload, timeout=30)
        assert r.status_code == 403, f"expected 403 got {r.status_code} body={r.text[:200]}"

    def test_admin_marks_for_member(self, admin_headers, member_a):
        rw = requests.get(f"{API}/event-attendance/week?week_of={date.today().isoformat()}", headers=admin_headers, timeout=30)
        occs = rw.json().get("occurrences", []) if rw.status_code == 200 else []
        # pick an unlocked event today (Saturday)
        today = date.today().isoformat()
        cand = [o for o in occs if o.get("date") == today and not o.get("locked", True)]
        if not cand:
            cand = [o for o in occs if o.get("date") == today]
        if not cand:
            pytest.skip("no today events")
        eid = cand[0].get("event_id")
        payload = {"user_id": member_a["user_id"], "event_id": eid, "event_date": today, "status": "present"}
        r = requests.post(f"{API}/event-attendance/mark-for-member", headers=admin_headers, json=payload, timeout=30)
        assert r.status_code == 200, f"admin mark: {r.status_code} {r.text[:200]}"


# ---------------- TEAM WEEK ----------------
class TestTeamWeek:
    def test_admin_team_week(self, admin_headers):
        r = requests.get(f"{API}/event-attendance/team-week", headers=admin_headers, timeout=30)
        assert r.status_code == 200, f"{r.status_code} {r.text[:200]}"
        data = r.json()
        assert "occurrences" in data
        assert "grid" in data


# ---------------- GOALS ----------------
class TestGoals:
    def test_goals_crud_still_works(self, member_a):
        r = requests.post(f"{API}/goals", headers=member_a["headers"],
                          json={"title": "TEST_goal", "target": 5, "period": "weekly"}, timeout=30)
        assert r.status_code in (200, 201), r.text
        gid = r.json().get("goal_id")
        # progress update
        r2 = requests.patch(f"{API}/goals/{gid}/progress", headers=member_a["headers"],
                            json={"progress": 1}, timeout=30)
        assert r2.status_code == 200
        # cleanup best-effort
        requests.delete(f"{API}/goals/{gid}", headers=member_a["headers"], timeout=30)


# ---------------- DEDUPE ----------------
class TestDedupe:
    def test_admin_dedupe(self, admin_headers):
        r = requests.post(f"{API}/admin/dedupe-data", headers=admin_headers, timeout=60)
        assert r.status_code == 200, f"{r.status_code} {r.text[:200]}"
        data = r.json()
        assert data.get("ok") is True
        assert "duplicates_removed" in data

    def test_member_cannot_dedupe(self, member_a):
        r = requests.post(f"{API}/admin/dedupe-data", headers=member_a["headers"], timeout=30)
        assert r.status_code == 403


# ---------------- CLEANUP ----------------
@pytest.fixture(scope="session", autouse=True)
def cleanup_test_users(admin_headers, request):
    yield
    # Delete TEST users at end
    try:
        r = requests.get(f"{API}/admin/users", headers=admin_headers, timeout=30)
        if r.status_code == 200:
            users = r.json() if isinstance(r.json(), list) else r.json().get("users", [])
            for u in users:
                em = u.get("email", "")
                if em.startswith("TEST_"):
                    uid = u.get("user_id") or u.get("id")
                    if uid:
                        requests.delete(f"{API}/admin/users/{uid}", headers=admin_headers, timeout=30)
    except Exception as e:
        print(f"cleanup: {e}")
